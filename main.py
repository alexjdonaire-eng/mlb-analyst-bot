print("MAIN FILE CARGADO")

import os
import requests

print("REQUESTS IMPORTADO")

from analyzer import analyze_games, fetch_mlb_games

print("ANALYZER IMPORTADO")

from tracker import save_pick, load_results
from supabase_client import supabase

print("TRACKER IMPORTADO")

from openpyxl import Workbook

print("OPENPYXL IMPORTADO")

from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from grader import grade_picks

print("GRADER IMPORTADO")

from openpyxl.formatting.rule import CellIsRule

print("TODO IMPORTADO")

TELEGRAM_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

EXCEL_PATH = "MIBOTMLB_Dashboard_Automatico.xlsx"

# ===========================================
# TELEGRAM
# ===========================================
def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    data = {
    "chat_id": CHAT_ID,
    "text": message
}
    requests.post(url, data=data)

# ===========================================
# FORMATO MENSAJE POR JUEGO
# ===========================================
def format_game(game):
    home = game.get("home_team")
    away = game.get("away_team")
    home_pitcher = game.get("home_pitcher", {"name": "TBD", "ERA": "-", "WHIP": "-"})
    away_pitcher = game.get("away_pitcher", {"name": "TBD", "ERA": "-", "WHIP": "-"})
    
    winner = game.get("predicted_winner", {})
    total = game.get("predicted_total", {})
    handicap = game.get("predicted_handicap", {})

    pick_type = game.get("top_pick_type", "Ganador")
    pick_value = game.get("top_pick_value", winner.get("team", ""))

    return (
f"⚾ {away} vs {home}\n\n"
f"🧾 Lanzadores\n"
f"{away}: {away_pitcher['name']} (ERA {away_pitcher['ERA']} | WHIP {away_pitcher['WHIP']})\n"
f"{home}: {home_pitcher['name']} (ERA {home_pitcher['ERA']} | WHIP {home_pitcher['WHIP']})\n\n"
f"🎯 Ganador: {winner.get('team', 'TBD')} ({winner.get('prob', 0)}%)\n"
f"⚾ Total: {total.get('type','-')} {total.get('line', '-')} ({total.get('prob', 0)}%)\n"
f"⚾ Hándicap: {handicap.get('line', '-')} ({handicap.get('prob', 0)}%)\n\n"
f"📊 Confianza: {game.get('confidence', 0)}%\n"
f"🏷 Nivel: {game.get('level', '🚫 PASAR')}\n"
f"💎 Pick: {pick_type} → {pick_value} ({game.get('confidence', 0)}%)"
    )

# ===========================================
# GENERAR DASHBOARD EXCEL
# ===========================================
def generate_excel(analyzed_games):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    today = datetime.now().strftime("%Y-%m-%d")
    results = load_results()
    # Título y fecha
    ws.merge_cells("A1:F1")
    ws["A1"] = "MIBOTMLB - Dashboard TOP 5"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    today_str = datetime.now().strftime("%Y-%m-%d")
    ws["A2"] = f"Fecha: {today_str}"
    ws["A2"].font = Font(bold=True)

    # Encabezados
    headers = ["Juego", "Mercado", "Pick", "Resultado", "Confianza", "Estado"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col)
        c.value = h
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAD3")

    # Solo TOP 5
    top5 = sorted(analyzed_games, key=lambda x: x.get("confidence",0), reverse=True)[:5]

    # Colores
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    row_start = 4
    for i, g in enumerate(top5, start=row_start):
        ws.cell(i, 1, g['top_pick_game'])
        ws.cell(i, 2, g['top_pick_type'])
        ws.cell(i, 3, g['top_pick_value'])
        # Cargar resultado desde tracker
        pick_result = "PENDIENTE"
        if today in results:
            for item in results[today]:
                if item["game"] == g['top_pick_game']:
                    pick_result = item["result"]
        ws.cell(i, 4, pick_result)
        ws.cell(i, 5, g.get("confidence",0))
        ws.cell(i, 6, "⏳" if pick_result=="PENDIENTE" else ("✅" if pick_result=="GANO" else "❌"))
        # Colores por resultado
        if pick_result=="GANO":
            ws.cell(i,6).fill = green_fill
        elif pick_result=="PERDIO":
            ws.cell(i,6).fill = red_fill
        else:
            ws.cell(i,6).fill = yellow_fill

    # Resumen al final
    ws["A12"] = "Ganados"
    ws["B12"] = sum(1 for g in top5 if load_results().get(today) and any(item["game"]==g['top_pick_game'] and item["result"]=="GANO" for item in load_results()[today]))
    ws["A13"] = "Perdidos"
    ws["B13"] = sum(1 for g in top5 if load_results().get(today) and any(item["game"]==g['top_pick_game'] and item["result"]=="PERDIO" for item in load_results()[today]))
    ws["A14"] = "Pendientes"
    ws["B14"] = len(top5) - ws["B12"].value - ws["B13"].value
    ws["A15"] = "Win Rate %"
    total = ws["B12"].value + ws["B13"].value
    ws["B15"] = round(ws["B12"].value / total * 100,1) if total>0 else 0

    # Semáforo Win Rate
    ws.conditional_formatting.add("B15", 
        CellIsRule(operator="lessThan", formula=['40'], fill=red_fill))
    ws.conditional_formatting.add("B15",
        CellIsRule(operator="between", formula=['40','55'], fill=yellow_fill))
    ws.conditional_formatting.add("B15",
        CellIsRule(operator="greaterThanOrEqual", formula=['55'], fill=green_fill))

    # Ajustar columnas
    from openpyxl.utils import get_column_letter
    for col in range(1,7):
        ws.column_dimensions[get_column_letter(col)].width = 28

    wb.save(EXCEL_PATH)
    return EXCEL_PATH

# ===========================================
# MAIN
# ===========================================
def main():

    print("📡 MIBOTMLB START")

    try:
        grade_picks()
    except Exception as e:
        print(f"Grader error: {e}")

    games = fetch_mlb_games()

    if not games:
        send_telegram_message(
            "⚠️ No hay juegos hoy o error al obtener datos."
        )
        return

    analyzed_games, top_message = analyze_games(games)

    # Enviar cada juego
    for g in analyzed_games:
        msg = format_game(g)
        send_telegram_message(msg)

    # Enviar TOP 5
    send_telegram_message(top_message)

    # Guardar picks TOP 5
    top5 = sorted(
        analyzed_games,
        key=lambda x: x.get("confidence", 0),
        reverse=True
    )[:5]

    for pick in top5:
        save_pick(
            pick["top_pick_game"],
            pick["top_pick_type"],
            pick["top_pick_value"],
            pick["confidence"],
            pick["level"]
        )

    # Generar Excel
    excel_file = generate_excel(analyzed_games)

    # Subir a Storage
    upload_dashboard_to_storage(excel_file)

    # Enviar Excel por Telegram
    send_telegram_file(excel_file)

    # Aviso final
    send_telegram_message("📊 Dashboard diario actualizado")

    print(f"Archivo Excel generado: {excel_file}")


def send_telegram_file(file_path):

    try:

        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"

        with open(file_path, "rb") as f:

            requests.post(
                url,
                data={"chat_id": CHAT_ID},
                files={"document": f}
            )

        print("✅ EXCEL ENVIADO A TELEGRAM")

    except Exception as e:

        print("❌ ERROR TELEGRAM FILE")
        print(e)


def upload_dashboard_to_storage(file_path):

    try:

        today = datetime.now().strftime("%Y-%m-%d")

        with open(file_path, "rb") as f:

            supabase.storage.from_("dashboard").upload(
                f"{today}.xlsx",
                f.read()
            )

        print("✅ EXCEL SUBIDO A STORAGE")

    except Exception as e:

        print("❌ ERROR STORAGE")
        print(e)


if __name__ == "__main__":
    main()
